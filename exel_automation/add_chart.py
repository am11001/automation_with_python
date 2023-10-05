from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference



wb = load_workbook('test.xlsx')
sheet = wb['Report']


min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


barchart = BarChart()

data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)

categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  

barchart.add_data(data, "B12")
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")

barchart.title = "Sales by productline"
barchart.style = 4
wb.save('barcharet_test.xlsx')